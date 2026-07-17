"""
NSE Nifty Indices - Historical Data Downloader (Streamlit)
For Arthashastra Finsec Providers Pvt Ltd (AFP)

Browser port of the CustomTkinter desktop app. Builds one .xlsx workbook per
run (one sheet per selected index, plus an optional combined sheet) entirely
in memory and serves it via a download button -- there is no server-side
output folder, since Streamlit Community Cloud's filesystem is ephemeral and
shared across every visitor.
"""

import io
from collections import defaultdict
from datetime import date, datetime, timedelta

import streamlit as st
from openpyxl import Workbook

from nse_client import (
    NSEClient,
    _to_num,
    _write_combined_sheet,
    _write_index_sheet,
    build_catalog,
    safe_sheet_name,
)

st.set_page_config(
    page_title="NSE Nifty Indices Downloader",
    page_icon="\U0001F4C8",
    layout="wide",
)

CATALOG_TTL_SECONDS = 24 * 60 * 60


@st.cache_resource
def get_client():
    return NSEClient()


@st.cache_data(ttl=CATALOG_TTL_SECONDS, show_spinner=False)
def get_catalog(_client):
    return build_catalog(_client, log_fn=lambda msg: None)


# ============================================================
#  Quick-set date range callbacks
#  (must run in on_click, BEFORE the date_input widgets below are
#  instantiated with the same session_state keys)
# ============================================================
def _set_prev_month():
    today = date.today()
    first_this = today.replace(day=1)
    last_prev = first_this - timedelta(days=1)
    st.session_state["start_date"] = last_prev.replace(day=1)
    st.session_state["end_date"] = last_prev


def _set_this_month():
    today = date.today()
    st.session_state["start_date"] = today.replace(day=1)
    st.session_state["end_date"] = today


def _set_ytd():
    today = date.today()
    st.session_state["start_date"] = date(today.year, 1, 1)
    st.session_state["end_date"] = today


def _set_last_year():
    today = date.today()
    st.session_state["start_date"] = today - timedelta(days=365)
    st.session_state["end_date"] = today


if "start_date" not in st.session_state:
    st.session_state["start_date"] = date.today() - timedelta(days=30)
if "end_date" not in st.session_state:
    st.session_state["end_date"] = date.today()


# ============================================================
#  Header
# ============================================================
st.title("NSE Nifty Indices — Historical Data Downloader")
st.caption("Arthashastra Finsec Providers Pvt Ltd (AFP)")

client = get_client()
with st.spinner("Loading index catalog from niftyindices.com ..."):
    catalog = get_catalog(client)

if not catalog or not any(catalog.values()):
    st.error(
        "Could not load the index catalog from niftyindices.com. "
        "The site may be temporarily unavailable -- try Refresh Catalog below, "
        "or check back later."
    )
    st.stop()


# ============================================================
#  Preset / selection callbacks
#  (defined after catalog load so they can close over it)
# ============================================================
def _clear_selection():
    for cat in catalog:
        st.session_state[f"cat_{cat}"] = []


def _save_preset():
    selected = set()
    for cat in catalog:
        selected.update(st.session_state.get(f"cat_{cat}", []))
    st.session_state["preset_indices"] = sorted(selected)
    st.session_state["_preset_msg"] = f"saved:{len(selected)}"


def _load_preset():
    preset = set(st.session_state.get("preset_indices", []))
    if not preset:
        st.session_state["_preset_msg"] = "empty"
        return
    for cat, indices in catalog.items():
        st.session_state[f"cat_{cat}"] = [i for i in indices if i in preset]
    st.session_state["_preset_msg"] = f"loaded:{len(preset)}"


def _refresh_catalog():
    get_catalog.clear()


left, right = st.columns([3, 2], gap="large")

with left:
    st.subheader("Select Indices")
    for cat in sorted(catalog):
        with st.expander(f"{cat}  ({len(catalog[cat])})"):
            st.multiselect(
                cat,
                options=catalog[cat],
                key=f"cat_{cat}",
                label_visibility="collapsed",
            )

    selected_indices = set()
    for cat in catalog:
        selected_indices.update(st.session_state.get(f"cat_{cat}", []))

    n_selected = len(selected_indices)
    st.caption("1 selected" if n_selected == 1 else f"{n_selected} selected")

    msg = st.session_state.pop("_preset_msg", None)
    if msg == "empty":
        st.info("No saved preset found. Select indices, then click 'Save Preset'.")
    elif msg and msg.startswith("saved:"):
        st.success(f"Preset saved ({msg.split(':')[1]} indices).")
    elif msg and msg.startswith("loaded:"):
        st.success(f"Loaded preset ({msg.split(':')[1]} indices).")

    b1, b2, b3, b4 = st.columns(4)
    b1.button("Save Preset", on_click=_save_preset, use_container_width=True)
    b2.button("Load Preset", on_click=_load_preset, use_container_width=True)
    b3.button("Clear", on_click=_clear_selection, use_container_width=True)
    b4.button("Refresh Catalog", on_click=_refresh_catalog, use_container_width=True)

with right:
    st.subheader("Configure Download")

    st.markdown("**Date range**")
    q1, q2, q3, q4 = st.columns(4)
    q1.button("Prev Month", on_click=_set_prev_month, use_container_width=True)
    q2.button("This Month", on_click=_set_this_month, use_container_width=True)
    q3.button("YTD", on_click=_set_ytd, use_container_width=True)
    q4.button("Last 1Y", on_click=_set_last_year, use_container_width=True)

    d1, d2 = st.columns(2)
    start_date = d1.date_input("Start", key="start_date")
    end_date = d2.date_input("End", key="end_date")

    st.markdown("**Options**")
    combined_sheet = st.checkbox(
        'Also include an "All Indices" combined sheet',
        value=True,
        key="combined_sheet",
    )
    st.caption(
        "Every download makes one .xlsx workbook for the date range you pick, "
        "with one sheet per selected index."
    )

    st.divider()

    download_clicked = st.button(
        "DOWNLOAD",
        type="primary",
        use_container_width=True,
        disabled=not selected_indices,
    )

    if download_clicked:
        if start_date > end_date:
            st.error("Start date must be on or before end date.")
        else:
            indices = sorted(selected_indices)
            n_total = len(indices)
            idx_rows = defaultdict(list)

            with st.status(f"Downloading {n_total} indices ...", expanded=True) as status:
                progress = st.progress(0.0)
                for i, idx in enumerate(indices, start=1):
                    st.write(f"[{i}/{n_total}] Fetching {idx} ...")
                    try:
                        rows = client.get_historical_data(idx, start_date, end_date)
                        for r in rows:
                            try:
                                d = datetime.strptime(r["HistoricalDate"], "%d %b %Y").date()
                            except (KeyError, ValueError):
                                continue
                            idx_rows[idx].append({
                                "index_name": r.get("INDEX_NAME", idx),
                                "date": d,
                                "open": _to_num(r.get("OPEN", "")),
                                "high": _to_num(r.get("HIGH", "")),
                                "low": _to_num(r.get("LOW", "")),
                                "close": _to_num(r.get("CLOSE", "")),
                            })
                        st.write(f"  + {len(rows)} records received")
                    except Exception as e:
                        st.write(f"  x ERROR: {e}")
                    progress.progress(i / n_total)

                st.write("Building workbook ...")
                wb = Workbook()
                wb.remove(wb.active)
                used_sheet_names = set()
                total_rows = 0

                for idx in indices:
                    rows = sorted(idx_rows.get(idx, []), key=lambda r: r["date"])
                    ws = wb.create_sheet(safe_sheet_name(idx, used_sheet_names))
                    _write_index_sheet(ws, rows)
                    total_rows += len(rows)
                    if not rows:
                        st.write(f"  ! {idx}: no data in range -- sheet left with headers only")

                if combined_sheet:
                    combined = []
                    for idx in indices:
                        combined.extend(idx_rows.get(idx, []))
                    combined.sort(key=lambda r: (r["date"], r["index_name"]))
                    ws = wb.create_sheet(safe_sheet_name("All Indices", used_sheet_names))
                    _write_combined_sheet(ws, combined)

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)

                fname = f"NSE_Indices_{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx"
                st.session_state["workbook_bytes"] = buf.getvalue()
                st.session_state["workbook_filename"] = fname
                st.session_state["workbook_summary"] = (
                    f"{total_rows} rows across {len(indices)} indices"
                )

                status.update(
                    label=f"Done — {total_rows} rows written to {fname}",
                    state="complete",
                )

    if st.session_state.get("workbook_bytes"):
        st.success(f"Ready: {st.session_state.get('workbook_summary', '')}")
        st.download_button(
            "Download workbook",
            data=st.session_state["workbook_bytes"],
            file_name=st.session_state["workbook_filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
