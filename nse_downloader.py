"""
NSE Nifty Indices  -  Historical Data Downloader  (Modern UI)
For Arthashastra Finsec Providers Pvt Ltd (AFP)

A standalone Windows desktop app built with CustomTkinter.

UI features
-----------
- Sepia palette  (warm parchment tones, fixed single theme)
- Custom collapsible index list with master / child checkboxes
- Live filter, animated progress bar, status indicator dot
- Cards with rounded corners, hover effects, modern typography

Functional features
-------------------
- Pre-select indices, save as default preset auto-loaded next launch
- One .xlsx workbook per download run, covering the full date range picked
  (no more monthly-split files): one sheet per selected index, plus an
  optional combined "All Indices" sheet
- Threaded downloads, responsive UI
- Local cache of the index catalog
"""

import json
import os
import queue
import re
import sys
import threading
from collections import defaultdict
from datetime import date, datetime, timedelta
from tkinter import filedialog, messagebox

import customtkinter as ctk
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
#  Sepia palette  -  warm parchment, single fixed theme
# ============================================================
P = {
    "bg":            "#F3E9D2",
    "surface":       "#FBF4E4",
    "surface_2":     "#EFDFBF",
    "surface_hover": "#E5D1A3",
    "border":        "#D6BE8E",
    "text":          "#4A3728",
    "text_muted":    "#8A7256",
    "primary":       "#8B5A2B",   # sepia brown
    "primary_hover": "#6E4720",
    "secondary":     "#B08442",   # amber / gold
    "success":       "#6B8E4E",
    "warning":       "#C08A2E",
    "error":         "#A24936",
}

# ============================================================
#  API constants
# ============================================================
BASE_URL = "https://www.niftyindices.com"
HEADERS = {
    "Content-Type": "application/json; charset=UTF-8",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "X-Requested-With": "XMLHttpRequest",
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer": "https://www.niftyindices.com/reports/historical-data",
    "Origin":  "https://www.niftyindices.com",
}

INDEX_TYPES = ["Equity", "Fixed Income", "Multi Asset"]
MAX_CHUNK_DAYS = 360


def app_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


CATALOG_PATH = os.path.join(app_dir(), "indices_catalog.json")
CONFIG_PATH  = os.path.join(app_dir(), "nse_downloader_config.json")


def safe_sheet_name(name, used_names):
    """Excel worksheet names: <=31 chars, no : \\ / ? * [ ], not blank,
    and must be unique within the workbook."""
    cleaned = re.sub(r"[:\\/?*\[\]]", "_", name).strip() or "Sheet"
    base = cleaned[:31]
    candidate = base
    n = 2
    while candidate.lower() in used_names:
        suffix = f" ({n})"
        candidate = base[: 31 - len(suffix)] + suffix
        n += 1
    used_names.add(candidate.lower())
    return candidate


def _to_num(v):
    """Convert an OHLC value to float where possible; leave untouched otherwise."""
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return v


_HEADER_FILL = PatternFill(start_color="E8EAF1", end_color="E8EAF1", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="1E1B2E")
_DATE_FMT = "yyyy-mm-dd"
_NUM_FMT = "#,##0.00"


def _style_header(ws, headers):
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"


def _write_index_sheet(ws, rows):
    """rows: list of dicts with date/open/high/low/close, already sorted by date."""
    _style_header(ws, ["Date", "Open", "High", "Low", "Close"])
    for r, row in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=row["date"]).number_format = _DATE_FMT
        for c, key in enumerate(("open", "high", "low", "close"), start=2):
            cell = ws.cell(row=r, column=c, value=row[key])
            if isinstance(row[key], (int, float)):
                cell.number_format = _NUM_FMT
    for i, w in enumerate([14, 13, 13, 13, 13], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_combined_sheet(ws, rows):
    """rows: list of dicts with index_name/date/open/high/low/close,
    already sorted by (date, index_name)."""
    _style_header(ws, ["Index", "Date", "Open", "High", "Low", "Close"])
    for r, row in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=row["index_name"])
        ws.cell(row=r, column=2, value=row["date"]).number_format = _DATE_FMT
        for c, key in enumerate(("open", "high", "low", "close"), start=3):
            cell = ws.cell(row=r, column=c, value=row[key])
            if isinstance(row[key], (int, float)):
                cell.number_format = _NUM_FMT
    for i, w in enumerate([22, 14, 13, 13, 13, 13], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def parse_date_str(s):
    s = (s or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


# ============================================================
#  API client
# ============================================================
class NSEClient:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        try:
            self.session.get(f"{BASE_URL}/reports/historical-data", timeout=20)
        except requests.RequestException:
            pass

    @staticmethod
    def _parse_json_response(r):
        """Parse a response as JSON, raising a clear error (instead of a
        cryptic JSONDecodeError) if the site ever serves back its HTML
        shell / a bot-check page instead of real JSON."""
        try:
            return r.json()
        except (json.JSONDecodeError, ValueError):
            snippet = r.text.strip().replace("\n", " ")[:120]
            raise RuntimeError(
                f"niftyindices.com returned non-JSON (status {r.status_code}): "
                f"{snippet!r} -- the site's API may have changed again, or "
                f"requests are being blocked as a bot."
            )

    def get_sub_index_types(self, index_type, index_group="Historical Index Data"):
        url = f"{BASE_URL}/BackPage/gethistoricaltypeSubindexdata"
        payload = {"cinfo": {"indextype": index_type, "indexgroup": index_group}}
        r = self.session.post(url, json=payload, timeout=30)
        r.raise_for_status()
        return [item["indextype"] for item in self._parse_json_response(r)]

    def get_indices(self, sub_index_type, index_group="Historical Index Data"):
        url = f"{BASE_URL}/BackPage/gethistoricaltypeindexdata"
        payload = {"cinfo": {"indextype": sub_index_type, "indexgroup": index_group}}
        r = self.session.post(url, json=payload, timeout=30)
        r.raise_for_status()
        return [item["indextype"] for item in self._parse_json_response(r)]

    def get_historical_data(self, index_name, start_date, end_date):
        all_records = []
        cursor = start_date
        while cursor <= end_date:
            chunk_end = min(cursor + timedelta(days=MAX_CHUNK_DAYS), end_date)
            all_records.extend(self._fetch_chunk(index_name, cursor, chunk_end))
            cursor = chunk_end + timedelta(days=1)
        seen, unique = set(), []
        for r in all_records:
            d = r.get("HistoricalDate", "")
            if d and d not in seen:
                seen.add(d)
                unique.append(r)
        unique.sort(key=lambda r: datetime.strptime(r["HistoricalDate"], "%d %b %Y"))
        return unique

    def _fetch_chunk(self, index_name, start_date, end_date):
        url = f"{BASE_URL}/BackPage/getHistoricaldatatabletoString"
        s = start_date.strftime("%d-%b-%Y")
        e = end_date.strftime("%d-%b-%Y")
        cinfo = (
            f"{{'name':'{index_name}','startDate':'{s}',"
            f"'endDate':'{e}','indexName':'{index_name}'}}"
        )
        r = self.session.post(url, json={"cinfo": cinfo}, timeout=60)
        r.raise_for_status()
        try:
            return self._parse_json_response(r)
        except RuntimeError:
            return []


def build_catalog(client, log_fn):
    catalog = {}
    for itype in INDEX_TYPES:
        log_fn(f"Fetching sub-types for: {itype}")
        try:
            subs = client.get_sub_index_types(itype)
        except Exception as e:
            log_fn(f"  x {itype}: {e}")
            continue
        for sub in subs:
            label = f"{sub} [{itype}]" if itype != "Equity" else sub
            try:
                idxs = client.get_indices(sub)
                catalog[label] = sorted(idxs)
                log_fn(f"  + {label}: {len(idxs)} indices")
            except Exception as e:
                log_fn(f"  x {label}: {e}")
                catalog[label] = []
    return catalog


# ============================================================
#  Custom collapsible index list with checkboxes
# ============================================================
class IndexListWidget(ctk.CTkFrame):

    def __init__(self, master, on_change=None, **kw):
        super().__init__(master, fg_color="transparent", **kw)
        self.on_change = on_change or (lambda: None)
        self.checked = set()
        self.categories = {}
        self._filter = ""

        # Filter row
        # Note: don't bind textvariable — CTk 5.2.2 placeholder hides itself
        # when a StringVar is bound. Use direct .get() + KeyRelease instead.
        self.filter_entry = ctk.CTkEntry(
            self,
            placeholder_text="Filter indices  (type to search)",
            font=ctk.CTkFont(size=13),
            border_width=1, border_color=P["border"],
            fg_color=P["surface"], text_color=P["text"],
            placeholder_text_color=P["text_muted"],
            corner_radius=8, height=36,
        )
        self.filter_entry.bind("<KeyRelease>", lambda e: self._on_filter_change())
        self.filter_entry.pack(fill="x", padx=2, pady=(0, 8))

        # Scrollable area
        self.scroll = ctk.CTkScrollableFrame(
            self,
            fg_color=P["surface"],
            corner_radius=10,
            border_width=1,
            border_color=P["border"],
            scrollbar_button_color=P["surface_2"],
            scrollbar_button_hover_color=P["primary"],
        )
        self.scroll.pack(fill="both", expand=True, padx=0, pady=0)

    # --- public API ---
    def set_catalog(self, catalog):
        for w in self.scroll.winfo_children():
            w.destroy()
        self.categories.clear()
        for cat_name in sorted(catalog.keys()):
            self._add_category(cat_name, catalog[cat_name])
        self._apply_filter()

    def get_checked(self):
        return set(self.checked)

    def set_checked(self, names_iter):
        names = set(names_iter)
        self.checked = set(names)
        for cat in self.categories.values():
            for ent in cat["children"]:
                if ent["name"] in names:
                    ent["var"].set("on")
                else:
                    ent["var"].set("off")
            self._sync_master(cat)
        self.on_change()

    def clear(self):
        self.set_checked([])

    # --- internal ---
    def _add_category(self, name, indices):
        container = ctk.CTkFrame(self.scroll, fg_color="transparent")
        container.pack(fill="x", pady=(2, 0))

        # Header pill
        header = ctk.CTkFrame(
            container, fg_color=P["surface_2"],
            corner_radius=8, height=38)
        header.pack(fill="x", padx=4, pady=2)
        header.pack_propagate(False)

        chevron = ctk.CTkLabel(
            header, text="\u25B6", width=16,
            font=ctk.CTkFont(size=11),
            text_color=P["text_muted"], cursor="hand2")
        chevron.pack(side="left", padx=(12, 4))

        master_var = ctk.StringVar(value="off")
        master_cb = ctk.CTkCheckBox(
            header, text=name, variable=master_var,
            onvalue="on", offvalue="off",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=P["text"],
            checkbox_width=18, checkbox_height=18,
            corner_radius=4, border_width=2,
            fg_color=P["primary"], hover_color=P["primary_hover"],
            border_color=P["text_muted"],
            command=lambda n=name: self._toggle_category(n),
        )
        master_cb.pack(side="left", padx=(2, 8))

        count_lbl = ctk.CTkLabel(
            header, text=f" {len(indices)} ",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=P["text_muted"],
            fg_color=P["surface"],
            corner_radius=8)
        count_lbl.pack(side="right", padx=12)

        # Children container (initially hidden)
        children_frame = ctk.CTkFrame(container, fg_color="transparent")

        children = []
        for idx_name in indices:
            var = ctk.StringVar(value="off")
            cb = ctk.CTkCheckBox(
                children_frame, text=idx_name, variable=var,
                onvalue="on", offvalue="off",
                font=ctk.CTkFont(size=12),
                text_color=P["text"],
                checkbox_width=16, checkbox_height=16,
                corner_radius=3, border_width=2,
                fg_color=P["primary"], hover_color=P["primary_hover"],
                border_color=P["text_muted"],
                command=lambda n=idx_name: self._toggle_index(n),
            )
            cb.pack(fill="x", padx=(40, 12), pady=2, anchor="w")
            children.append({"name": idx_name, "cb": cb, "var": var})

        cat = {
            "container": container, "header": header, "chevron": chevron,
            "master_cb": master_cb, "master_var": master_var,
            "children_frame": children_frame, "children": children,
            "indices": list(indices), "expanded": False,
        }
        self.categories[name] = cat

        chevron.bind("<Button-1>", lambda e, n=name: self._toggle_expand(n))

    def _toggle_expand(self, cat_name):
        cat = self.categories[cat_name]
        cat["expanded"] = not cat["expanded"]
        cat["chevron"].configure(text="\u25BC" if cat["expanded"] else "\u25B6")
        if cat["expanded"]:
            cat["children_frame"].pack(fill="x", padx=4, pady=(0, 4))
        else:
            cat["children_frame"].pack_forget()

    def _toggle_index(self, idx_name):
        for cat in self.categories.values():
            for ent in cat["children"]:
                if ent["name"] == idx_name:
                    if ent["var"].get() == "on":
                        self.checked.add(idx_name)
                    else:
                        self.checked.discard(idx_name)
                    self._sync_master(cat)
                    self.on_change()
                    return

    def _toggle_category(self, cat_name):
        cat = self.categories[cat_name]
        new_state = (cat["master_var"].get() == "on")
        for ent in cat["children"]:
            if new_state:
                ent["var"].set("on")
                self.checked.add(ent["name"])
            else:
                ent["var"].set("off")
                self.checked.discard(ent["name"])
        self.on_change()

    def _sync_master(self, cat):
        all_on = bool(cat["indices"]) and all(n in self.checked for n in cat["indices"])
        cat["master_var"].set("on" if all_on else "off")

    def _on_filter_change(self):
        self._filter = self.filter_entry.get().strip().lower()
        self._apply_filter()

    def _apply_filter(self):
        q = self._filter
        for name, cat in self.categories.items():
            matching = [e for e in cat["children"] if not q or q in e["name"].lower()]
            if not matching:
                cat["container"].pack_forget()
                continue
            cat["container"].pack(fill="x", pady=(2, 0))
            if q:
                for ent in cat["children"]:
                    if q in ent["name"].lower():
                        ent["cb"].pack(fill="x", padx=(40, 12), pady=2, anchor="w")
                    else:
                        ent["cb"].pack_forget()
                if not cat["expanded"]:
                    cat["expanded"] = True
                    cat["chevron"].configure(text="\u25BC")
                    cat["children_frame"].pack(fill="x", padx=4, pady=(0, 4))
            else:
                for ent in cat["children"]:
                    ent["cb"].pack(fill="x", padx=(40, 12), pady=2, anchor="w")
                if cat["expanded"]:
                    cat["expanded"] = False
                    cat["chevron"].configure(text="\u25B6")
                    cat["children_frame"].pack_forget()


# ============================================================
#  Main application window
# ============================================================
class NSEDownloaderApp(ctk.CTk):

    def __init__(self):
        super().__init__()
        self.title("NSE Nifty Indices Downloader  -  AFP")
        self.geometry("1300x820")
        self.minsize(1080, 640)
        self.configure(fg_color=P["bg"])

        self.client = NSEClient()
        self.catalog = {}
        self.config = self._load_config()
        self.msg_queue = queue.Queue()
        self.busy = False

        ctk.set_appearance_mode("Light")

        self._build_ui()
        self._poll_queue()
        self._init_catalog()

    # ---------- UI build ----------
    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self._build_header()

        body = ctk.CTkFrame(self, fg_color="transparent")
        body.grid(row=1, column=0, sticky="nsew", padx=20, pady=(0, 8))
        body.grid_columnconfigure(0, weight=3, uniform="cols")
        body.grid_columnconfigure(1, weight=2, uniform="cols")
        body.grid_rowconfigure(0, weight=1)

        self._build_indices_card(body)
        self._build_options_card(body)
        self._build_action_bar()
        self._build_log_panel()

    # --- header ---
    def _build_header(self):
        header = ctk.CTkFrame(self, fg_color="transparent", height=72)
        header.grid(row=0, column=0, sticky="ew", padx=20, pady=(16, 12))
        header.grid_columnconfigure(1, weight=1)

        brand = ctk.CTkFrame(header, fg_color="transparent")
        brand.grid(row=0, column=0, sticky="w")

        ctk.CTkLabel(
            brand, text="\u25C6", font=ctk.CTkFont(size=24),
            text_color=P["primary"], width=36).pack(side="left", padx=(0, 8))

        title_box = ctk.CTkFrame(brand, fg_color="transparent")
        title_box.pack(side="left")
        ctk.CTkLabel(
            title_box, text="NSE Nifty Indices",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=P["text"]).pack(anchor="w")
        ctk.CTkLabel(
            title_box, text="Historical Data Downloader",
            font=ctk.CTkFont(size=11),
            text_color=P["text_muted"]).pack(anchor="w")

    # --- indices card ---
    def _build_indices_card(self, parent):
        card = ctk.CTkFrame(
            parent, fg_color=P["surface"],
            corner_radius=14, border_width=1,
            border_color=P["border"])
        card.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        card.grid_columnconfigure(0, weight=1)
        card.grid_rowconfigure(2, weight=1)

        title_row = ctk.CTkFrame(card, fg_color="transparent")
        title_row.grid(row=0, column=0, sticky="ew", padx=20, pady=(16, 8))
        title_row.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            title_row, text="Select Indices",
            font=ctk.CTkFont(size=15, weight="bold"),
            text_color=P["text"]).grid(row=0, column=0, sticky="w")

        self.selection_count_var = ctk.StringVar(value="0 selected")
        self.selection_pill = ctk.CTkLabel(
            title_row, textvariable=self.selection_count_var,
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=("#ffffff", "#ffffff"),
            fg_color=P["primary"], corner_radius=10)
        self.selection_pill.grid(row=0, column=1, sticky="e", ipadx=10, ipady=2)

        list_wrap = ctk.CTkFrame(card, fg_color="transparent")
        list_wrap.grid(row=2, column=0, sticky="nsew", padx=16, pady=(4, 8))
        list_wrap.grid_columnconfigure(0, weight=1)
        list_wrap.grid_rowconfigure(0, weight=1)

        self.index_list = IndexListWidget(
            list_wrap, on_change=self._on_selection_change)
        self.index_list.grid(row=0, column=0, sticky="nsew")

        action_row = ctk.CTkFrame(card, fg_color="transparent")
        action_row.grid(row=3, column=0, sticky="ew", padx=20, pady=(8, 16))
        action_row.grid_columnconfigure(3, weight=1)

        for col, (label, cmd) in enumerate([
            ("Save Preset",  self._save_preset),
            ("Load Preset",  self._load_preset),
            ("Clear",        self._clear_selection),
        ]):
            ctk.CTkButton(
                action_row, text=label, command=cmd,
                width=110, height=32, corner_radius=8,
                font=ctk.CTkFont(size=12, weight="bold"),
                fg_color=P["surface_2"], hover_color=P["surface_hover"],
                text_color=P["text"], border_width=0,
            ).grid(row=0, column=col, padx=(0 if col == 0 else 6, 0))

        ctk.CTkButton(
            action_row, text="Refresh Catalog",
            command=self._refresh_catalog,
            width=130, height=32, corner_radius=8,
            font=ctk.CTkFont(size=12),
            fg_color="transparent", hover_color=P["surface_2"],
            text_color=P["text_muted"], border_width=1,
            border_color=P["border"],
        ).grid(row=0, column=4, padx=(8, 0), sticky="e")

    # --- options card ---
    def _build_options_card(self, parent):
        card = ctk.CTkFrame(
            parent, fg_color=P["surface"],
            corner_radius=14, border_width=1,
            border_color=P["border"])
        card.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        card.grid_columnconfigure(0, weight=1)
        card.grid_rowconfigure(1, weight=1)

        # Title pinned at top
        ctk.CTkLabel(
            card, text="Configure Download",
            font=ctk.CTkFont(size=15, weight="bold"),
            text_color=P["text"]).grid(
            row=0, column=0, sticky="w", padx=20, pady=(16, 4))

        # Scrollable inner area — sections go in here so nothing is ever
        # clipped at small window sizes or high DPI scaling
        inner = ctk.CTkScrollableFrame(
            card,
            fg_color="transparent",
            scrollbar_button_color=P["surface_2"],
            scrollbar_button_hover_color=P["primary"],
        )
        inner.grid(row=1, column=0, sticky="nsew", padx=8, pady=(0, 12))
        inner.grid_columnconfigure(0, weight=1)

        # --- Date range ---
        self._section_label(inner, "DATE RANGE", row=1)
        date_box = ctk.CTkFrame(inner, fg_color="transparent")
        date_box.grid(row=2, column=0, sticky="ew", padx=12, pady=(0, 4))
        date_box.grid_columnconfigure(0, weight=1)
        date_box.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            date_box, text="Start", font=ctk.CTkFont(size=11),
            text_color=P["text_muted"]).grid(
            row=0, column=0, sticky="w", padx=(0, 6))
        ctk.CTkLabel(
            date_box, text="End", font=ctk.CTkFont(size=11),
            text_color=P["text_muted"]).grid(
            row=0, column=1, sticky="w", padx=(6, 0))

        self.start_entry = ctk.CTkEntry(
            date_box,
            placeholder_text="YYYY-MM-DD",
            font=ctk.CTkFont(size=13),
            border_width=1, border_color=P["border"],
            fg_color=P["surface"], text_color=P["text"],
            placeholder_text_color=P["text_muted"],
            corner_radius=8, height=34,
        )
        self.start_entry.grid(row=1, column=0, sticky="ew", padx=(0, 6), pady=(4, 0))

        self.end_entry = ctk.CTkEntry(
            date_box,
            placeholder_text="YYYY-MM-DD",
            font=ctk.CTkFont(size=13),
            border_width=1, border_color=P["border"],
            fg_color=P["surface"], text_color=P["text"],
            placeholder_text_color=P["text_muted"],
            corner_radius=8, height=34,
        )
        self.end_entry.grid(row=1, column=1, sticky="ew", padx=(6, 0), pady=(4, 0))

        quick = ctk.CTkFrame(inner, fg_color="transparent")
        quick.grid(row=3, column=0, sticky="ew", padx=12, pady=(8, 4))
        for col, (label, fn) in enumerate([
            ("Prev Month", self._set_prev_month),
            ("This Month", self._set_this_month),
            ("YTD",        self._set_ytd),
            ("Last 1Y",    self._set_last_year),
        ]):
            ctk.CTkButton(
                quick, text=label, command=fn,
                width=80, height=28, corner_radius=6,
                font=ctk.CTkFont(size=11),
                fg_color="transparent", hover_color=P["surface_2"],
                text_color=P["secondary"], border_width=1,
                border_color=P["border"],
            ).grid(row=0, column=col, padx=(0 if col == 0 else 6, 0), sticky="ew")
            quick.grid_columnconfigure(col, weight=1)

        # --- Output folder ---
        self._section_label(inner, "OUTPUT FOLDER", row=4)
        out_box = ctk.CTkFrame(inner, fg_color="transparent")
        out_box.grid(row=5, column=0, sticky="ew", padx=12, pady=(0, 4))
        out_box.grid_columnconfigure(0, weight=1)

        default_outdir = self.config.get(
            "output_folder", os.path.join(app_dir(), "NSE_Data"))
        self.outdir_entry = ctk.CTkEntry(
            out_box,
            placeholder_text="C:\\path\\to\\output\\folder",
            font=ctk.CTkFont(size=12),
            border_width=1, border_color=P["border"],
            fg_color=P["surface"], text_color=P["text"],
            placeholder_text_color=P["text_muted"],
            corner_radius=8, height=34,
        )
        if default_outdir:
            self.outdir_entry.insert(0, default_outdir)
        self.outdir_entry.grid(row=0, column=0, sticky="ew", padx=(0, 6))

        ctk.CTkButton(
            out_box, text="Browse",
            command=self._browse_outdir,
            width=80, height=34, corner_radius=8,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color=P["surface_2"], hover_color=P["surface_hover"],
            text_color=P["text"], border_width=0,
        ).grid(row=0, column=1)

        # --- Behavior options ---
        self._section_label(inner, "OPTIONS", row=7)

        opt_box = ctk.CTkFrame(inner, fg_color="transparent")
        opt_box.grid(row=8, column=0, sticky="ew", padx=12, pady=(0, 4))

        self.combined_sheet_var = ctk.BooleanVar(
            value=self.config.get("combined_sheet", True))

        ctk.CTkCheckBox(
            opt_box,
            text="Also include an \"All Indices\" combined sheet",
            variable=self.combined_sheet_var,
            font=ctk.CTkFont(size=12),
            text_color=P["text"],
            checkbox_width=18, checkbox_height=18,
            corner_radius=4, border_width=2,
            fg_color=P["primary"], hover_color=P["primary_hover"],
            border_color=P["text_muted"],
        ).pack(fill="x", anchor="w", pady=3)

        ctk.CTkLabel(
            opt_box,
            text=("Every download makes one .xlsx workbook for the date "
                  "range you pick, with one sheet per selected index."),
            font=ctk.CTkFont(size=11),
            text_color=P["text_muted"],
            wraplength=260, justify="left",
        ).pack(fill="x", anchor="w", pady=(4, 0))

    def _build_action_bar(self):
        """Sticky bottom action bar: status + progress + DOWNLOAD."""
        bar = ctk.CTkFrame(
            self, fg_color=P["surface"],
            corner_radius=14, border_width=1,
            border_color=P["border"], height=78)
        bar.grid(row=2, column=0, sticky="ew", padx=20, pady=(0, 8))
        bar.grid_propagate(False)
        bar.grid_columnconfigure(1, weight=1)

        # --- status block ---
        status_box = ctk.CTkFrame(bar, fg_color="transparent")
        status_box.grid(row=0, column=0, sticky="w", padx=(20, 14), pady=18)

        self.status_dot = ctk.CTkLabel(
            status_box, text="\u25CF", font=ctk.CTkFont(size=16),
            text_color=P["success"], width=18)
        self.status_dot.pack(side="left", padx=(0, 8))

        self.status_var = ctk.StringVar(value="Ready")
        ctk.CTkLabel(
            status_box, textvariable=self.status_var,
            font=ctk.CTkFont(size=13),
            text_color=P["text"]).pack(side="left")

        # --- progress bar ---
        self.progress = ctk.CTkProgressBar(
            bar, mode="determinate", height=8,
            corner_radius=4,
            progress_color=P["primary"],
            fg_color=P["surface_2"])
        self.progress.set(0)
        self.progress.grid(row=0, column=1, sticky="ew", padx=14, pady=24)

        # --- DOWNLOAD button ---
        self.download_btn = ctk.CTkButton(
            bar, text="DOWNLOAD",
            command=self._start_download,
            height=44, width=200, corner_radius=10,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=P["primary"],
            hover_color=P["primary_hover"],
            text_color=("#ffffff", "#ffffff"))
        self.download_btn.grid(row=0, column=2, sticky="e", padx=(14, 16), pady=17)

    def _section_label(self, parent, text, row):
        ctk.CTkLabel(
            parent, text=text,
            font=ctk.CTkFont(size=10, weight="bold"),
            text_color=P["text_muted"], anchor="w",
        ).grid(row=row, column=0, sticky="w", padx=12, pady=(10, 2))

    # --- log panel ---
    def _build_log_panel(self):
        log_card = ctk.CTkFrame(
            self, fg_color=P["surface"],
            corner_radius=14, border_width=1,
            border_color=P["border"], height=150)
        log_card.grid(row=3, column=0, sticky="ew", padx=20, pady=(0, 16))
        log_card.grid_columnconfigure(0, weight=1)
        log_card.grid_rowconfigure(1, weight=1)
        log_card.grid_propagate(False)

        title_row = ctk.CTkFrame(log_card, fg_color="transparent")
        title_row.grid(row=0, column=0, sticky="ew", padx=20, pady=(12, 4))
        title_row.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            title_row, text="Activity",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=P["text"]).grid(row=0, column=0, sticky="w")

        ctk.CTkButton(
            title_row, text="Clear", command=self._clear_log,
            width=70, height=24, corner_radius=6,
            font=ctk.CTkFont(size=11),
            fg_color="transparent", hover_color=P["surface_2"],
            text_color=P["text_muted"], border_width=1,
            border_color=P["border"],
        ).grid(row=0, column=1, sticky="e")

        self.log_box = ctk.CTkTextbox(
            log_card,
            font=ctk.CTkFont(family="Consolas", size=11),
            fg_color=P["surface_2"],
            text_color=P["text"],
            corner_radius=8,
            border_width=0,
            scrollbar_button_color=P["surface"],
            scrollbar_button_hover_color=P["primary"],
        )
        self.log_box.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 14))
        self.log_box.configure(state="disabled")

    # ---------- catalog management ----------
    def _init_catalog(self):
        if os.path.exists(CATALOG_PATH):
            try:
                with open(CATALOG_PATH, "r", encoding="utf-8") as f:
                    self.catalog = json.load(f)
                self.index_list.set_catalog(self.catalog)
                self._apply_saved_preset()
                total = sum(len(v) for v in self.catalog.values())
                self._log(f"Loaded {total} indices from local cache.")
                self._set_status("Ready", "success")
                return
            except Exception:
                self._log("Catalog cache invalid - re-fetching...")
        self._refresh_catalog()

    def _refresh_catalog(self):
        if self.busy:
            return
        self._set_busy(True, "Loading catalog...", "working")
        self._log("Refreshing index catalog from niftyindices.com ...")
        threading.Thread(target=self._refresh_catalog_thread, daemon=True).start()

    def _refresh_catalog_thread(self):
        try:
            cat = build_catalog(self.client, self._log_async)
            with open(CATALOG_PATH, "w", encoding="utf-8") as f:
                json.dump(cat, f, indent=2)
            self.msg_queue.put(("catalog_loaded", cat))
        except Exception as e:
            self.msg_queue.put(("catalog_error", str(e)))

    # ---------- selection / preset ----------
    def _on_selection_change(self):
        n = len(self.index_list.get_checked())
        self.selection_count_var.set(
            f"{n} selected" if n != 1 else "1 selected")

    def _save_preset(self):
        if self._save_config():
            n = len(self.index_list.get_checked())
            self._set_status(f"Preset saved  ({n} indices)", "success")
            self._log(f"Preset saved with {n} indices.")

    def _load_preset(self):
        preset = self.config.get("preset_indices", [])
        if not preset:
            messagebox.showinfo(
                "No Preset",
                "No saved preset found.\n\n"
                "Select indices, then click 'Save Preset'.")
            return
        self.index_list.set_checked(preset)
        n = len(self.index_list.get_checked())
        self._set_status(f"Loaded preset  ({n} indices)", "success")
        self._log(f"Loaded preset ({n} indices).")

    def _apply_saved_preset(self):
        preset = self.config.get("preset_indices", [])
        if preset:
            self.index_list.set_checked(preset)

    def _clear_selection(self):
        self.index_list.clear()

    def _set_entry_text(self, entry, value):
        entry.delete(0, "end")
        if value:
            entry.insert(0, value)

    # ---------- date helpers ----------
    def _set_prev_month(self):
        today = date.today()
        first_this = today.replace(day=1)
        last_prev = first_this - timedelta(days=1)
        first_prev = last_prev.replace(day=1)
        self._set_entry_text(self.start_entry, first_prev.isoformat())
        self._set_entry_text(self.end_entry, last_prev.isoformat())

    def _set_this_month(self):
        today = date.today()
        self._set_entry_text(self.start_entry, today.replace(day=1).isoformat())
        self._set_entry_text(self.end_entry, today.isoformat())

    def _set_ytd(self):
        today = date.today()
        self._set_entry_text(self.start_entry, date(today.year, 1, 1).isoformat())
        self._set_entry_text(self.end_entry, today.isoformat())

    def _set_last_year(self):
        today = date.today()
        self._set_entry_text(self.start_entry, (today - timedelta(days=365)).isoformat())
        self._set_entry_text(self.end_entry, today.isoformat())

    def _browse_outdir(self):
        d = filedialog.askdirectory(
            initialdir=self.outdir_entry.get() or app_dir(),
            title="Select output folder")
        if d:
            self._set_entry_text(self.outdir_entry, d)

    # ---------- config ----------
    def _load_config(self):
        if os.path.exists(CONFIG_PATH):
            try:
                with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return {}

    def _save_config(self):
        cfg = {
            "preset_indices": (sorted(self.index_list.get_checked())
                               if hasattr(self, "index_list")
                               else self.config.get("preset_indices", [])),
            "output_folder":  self.outdir_entry.get(),
            "combined_sheet": bool(self.combined_sheet_var.get()),
        }
        try:
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(cfg, f, indent=2)
            self.config = cfg
            return True
        except Exception as e:
            self._log(f"Could not save config: {e}")
            return False

    def _save_config_silent(self):
        try:
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(self.config, f, indent=2)
        except Exception:
            pass

    # ---------- download workflow ----------
    def _start_download(self):
        if self.busy:
            return
        checked = self.index_list.get_checked()
        if not checked:
            messagebox.showwarning(
                "No Selection",
                "Please select at least one index to download.")
            return
        start = parse_date_str(self.start_entry.get())
        end   = parse_date_str(self.end_entry.get())
        if not start or not end:
            messagebox.showerror(
                "Invalid Dates",
                "Please enter valid start and end dates in YYYY-MM-DD format.")
            return
        if start > end:
            messagebox.showerror(
                "Invalid Range",
                "Start date must be on or before end date.")
            return
        outdir = self.outdir_entry.get().strip()
        if not outdir:
            messagebox.showerror("No Output Folder",
                                 "Please choose an output folder.")
            return
        try:
            os.makedirs(outdir, exist_ok=True)
        except Exception as e:
            messagebox.showerror("Folder Error",
                                 f"Cannot create folder:\n{e}")
            return
        self._save_config()
        self._set_busy(True, "Downloading...", "working")
        self.progress.configure(mode="determinate")
        self.progress.set(0)

        opts = {
            "start": start, "end": end, "outdir": outdir,
            "indices": sorted(checked),
            "combined_sheet": self.combined_sheet_var.get(),
        }
        threading.Thread(
            target=self._download_thread, args=(opts,), daemon=True).start()

    def _download_thread(self, opts):
        try:
            indices = opts["indices"]
            start, end = opts["start"], opts["end"]
            outdir = opts["outdir"]
            self._log_async("")
            self._log_async("=== DOWNLOAD STARTED ===")
            self._log_async(f"Range:   {start.isoformat()}  ->  {end.isoformat()}")
            self._log_async(f"Indices: {len(indices)}")
            self._log_async(f"Output:  {outdir}")

            n_total = len(indices)
            idx_rows = defaultdict(list)
            for i, idx in enumerate(indices, 1):
                self._status_async(f"[{i}/{n_total}]  {idx}", "working")
                self._log_async("")
                self._log_async(f"[{i}/{n_total}] Fetching {idx} ...")
                try:
                    rows = self.client.get_historical_data(idx, start, end)
                    self._log_async(f"  + {len(rows)} records received")
                    for r in rows:
                        try:
                            d = datetime.strptime(
                                r["HistoricalDate"], "%d %b %Y").date()
                        except (KeyError, ValueError):
                            continue
                        idx_rows[idx].append({
                            "index_name": r.get("INDEX_NAME", idx),
                            "date":       d,
                            "open":       _to_num(r.get("OPEN", "")),
                            "high":       _to_num(r.get("HIGH", "")),
                            "low":        _to_num(r.get("LOW", "")),
                            "close":      _to_num(r.get("CLOSE", "")),
                        })
                except Exception as e:
                    self._log_async(f"  x ERROR: {e}")
                self.msg_queue.put(("progress", i / n_total))

            self._log_async("")
            self._log_async("Building workbook...")

            fname = f"NSE_Indices_{start.isoformat()}_to_{end.isoformat()}.xlsx"
            fpath = os.path.join(outdir, fname)
            if os.path.exists(fpath):
                self._log_async(f"  ! {fpath} already exists -- it will be overwritten")

            wb = Workbook()
            wb.remove(wb.active)  # drop the default blank sheet
            used_sheet_names = set()
            total_rows = 0

            for idx in indices:
                rows = sorted(idx_rows.get(idx, []), key=lambda r: r["date"])
                sheet_name = safe_sheet_name(idx, used_sheet_names)
                ws = wb.create_sheet(sheet_name)
                _write_index_sheet(ws, rows)
                total_rows += len(rows)
                if not rows:
                    self._log_async(f"  ! {idx}: no data in range -- sheet left with headers only")

            if opts["combined_sheet"]:
                combined = []
                for idx in indices:
                    combined.extend(idx_rows.get(idx, []))
                combined.sort(key=lambda r: (r["date"], r["index_name"]))
                ws = wb.create_sheet(safe_sheet_name("All Indices", used_sheet_names))
                _write_combined_sheet(ws, combined)

            try:
                wb.save(fpath)
            except PermissionError:
                raise RuntimeError(
                    f"Could not save {fpath} -- close it in Excel first, then try again.")

            self._log_async(f"  + {fpath}  ({len(indices)} sheets, {total_rows} rows)")
            self._log_async("")
            self._log_async(f"=== DONE ===   {total_rows} rows written to {fname}")
            self.msg_queue.put((
                "download_done",
                (f"Done  -  {fname}  ({total_rows} rows)", "success")))
        except Exception as e:
            self._log_async("")
            self._log_async(f"x Download failed: {e}")
            self.msg_queue.put((
                "download_done", (f"Failed: {e}", "error")))

    # ---------- thread <-> UI ----------
    def _log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_box.configure(state="normal")
        self.log_box.insert("end", f"[{ts}]  {msg}\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def _log_async(self, msg):
        self.msg_queue.put(("log", msg))

    def _status_async(self, msg, kind="working"):
        self.msg_queue.put(("status", (msg, kind)))

    def _set_status(self, msg, kind="ready"):
        self.status_var.set(msg)
        color_map = {
            "ready":   P["success"],
            "success": P["success"],
            "working": P["primary"],
            "error":   P["error"],
            "warning": P["warning"],
            "muted":   P["text_muted"],
        }
        self.status_dot.configure(text_color=color_map.get(kind, P["text_muted"]))

    def _poll_queue(self):
        try:
            while True:
                kind, payload = self.msg_queue.get_nowait()
                if kind == "log":
                    self._log(payload)
                elif kind == "status":
                    msg, status_kind = payload
                    self._set_status(msg, status_kind)
                elif kind == "progress":
                    self.progress.set(payload)
                elif kind == "catalog_loaded":
                    self.catalog = payload
                    self.index_list.set_catalog(payload)
                    self._apply_saved_preset()
                    total = sum(len(v) for v in payload.values())
                    self._set_busy(False, f"Loaded {total} indices.", "success")
                elif kind == "catalog_error":
                    self._set_busy(False, "Failed to load catalog.", "error")
                    messagebox.showerror(
                        "Catalog Error",
                        f"Could not fetch catalog:\n{payload}")
                elif kind == "download_done":
                    msg, kind2 = payload
                    self._set_busy(False, msg, kind2)
                    self.progress.set(1.0 if kind2 == "success" else 0)
        except queue.Empty:
            pass
        self.after(120, self._poll_queue)

    def _set_busy(self, busy, status=None, status_kind="ready"):
        self.busy = busy
        if status:
            self._set_status(status, status_kind)
        if busy:
            self.download_btn.configure(state="disabled")
        else:
            self.download_btn.configure(state="normal")


def main():
    app = NSEDownloaderApp()
    app.mainloop()


if __name__ == "__main__":
    main()
