"""
NSE Nifty Indices - API client and workbook-writing helpers.
For Arthashastra Finsec Providers Pvt Ltd (AFP)

Pure requests + openpyxl, no UI framework dependency. Ported from the
CustomTkinter desktop app (nse_downloader.py) for reuse in the Streamlit app.
"""

import json
import re
from datetime import datetime, timedelta

import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
#  API constants
# ============================================================
BASE_URL = "https://www.niftyindices.com"
HEADERS = {
    "Content-Type": "application/json; charset=UTF-8",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "X-Requested-With": "XMLHttpRequest",
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer": "https://www.niftyindices.com/reports/historical-data",
    "Origin":  "https://www.niftyindices.com",
}

INDEX_TYPES = ["Equity", "Fixed Income", "Multi Asset"]
MAX_CHUNK_DAYS = 360


# ============================================================
#  API client
# ============================================================
class NSEClient:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        try:
            self.session.get(f"{BASE_URL}/reports/historical-data", timeout=20)
        except requests.RequestException:
            pass

    @staticmethod
    def _parse_json_response(r):
        """Parse a response as JSON, raising a clear error (instead of a
        cryptic JSONDecodeError) if the site ever serves back its HTML
        shell / a bot-check page instead of real JSON."""
        try:
            return r.json()
        except (json.JSONDecodeError, ValueError):
            snippet = r.text.strip().replace("\n", " ")[:120]
            raise RuntimeError(
                f"niftyindices.com returned non-JSON (status {r.status_code}): "
                f"{snippet!r} -- the site's API may have changed again, or "
                f"requests are being blocked as a bot."
            )

    def get_sub_index_types(self, index_type, index_group="Historical Index Data"):
        url = f"{BASE_URL}/BackPage/gethistoricaltypeSubindexdata"
        payload = {"cinfo": {"indextype": index_type, "indexgroup": index_group}}
        r = self.session.post(url, json=payload, timeout=30)
        r.raise_for_status()
        return [item["indextype"] for item in self._parse_json_response(r)]

    def get_indices(self, sub_index_type, index_group="Historical Index Data"):
        url = f"{BASE_URL}/BackPage/gethistoricaltypeindexdata"
        payload = {"cinfo": {"indextype": sub_index_type, "indexgroup": index_group}}
        r = self.session.post(url, json=payload, timeout=30)
        r.raise_for_status()
        return [item["indextype"] for item in self._parse_json_response(r)]

    def get_historical_data(self, index_name, start_date, end_date):
        all_records = []
        cursor = start_date
        while cursor <= end_date:
            chunk_end = min(cursor + timedelta(days=MAX_CHUNK_DAYS), end_date)
            all_records.extend(self._fetch_chunk(index_name, cursor, chunk_end))
            cursor = chunk_end + timedelta(days=1)
        seen, unique = set(), []
        for r in all_records:
            d = r.get("HistoricalDate", "")
            if d and d not in seen:
                seen.add(d)
                unique.append(r)
        unique.sort(key=lambda r: datetime.strptime(r["HistoricalDate"], "%d %b %Y"))
        return unique

    def _fetch_chunk(self, index_name, start_date, end_date):
        url = f"{BASE_URL}/BackPage/getHistoricaldatatabletoString"
        s = start_date.strftime("%d-%b-%Y")
        e = end_date.strftime("%d-%b-%Y")
        cinfo = (
            f"{{'name':'{index_name}','startDate':'{s}',"
            f"'endDate':'{e}','indexName':'{index_name}'}}"
        )
        r = self.session.post(url, json={"cinfo": cinfo}, timeout=60)
        r.raise_for_status()
        try:
            return self._parse_json_response(r)
        except RuntimeError:
            return []


def build_catalog(client, log_fn):
    catalog = {}
    for itype in INDEX_TYPES:
        log_fn(f"Fetching sub-types for: {itype}")
        try:
            subs = client.get_sub_index_types(itype)
        except Exception as e:
            log_fn(f"  x {itype}: {e}")
            continue
        for sub in subs:
            label = f"{sub} [{itype}]" if itype != "Equity" else sub
            try:
                idxs = client.get_indices(sub)
                catalog[label] = sorted(idxs)
                log_fn(f"  + {label}: {len(idxs)} indices")
            except Exception as e:
                log_fn(f"  x {label}: {e}")
                catalog[label] = []
    return catalog


# ============================================================
#  Date parsing
# ============================================================
def parse_date_str(s):
    s = (s or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


# ============================================================
#  Workbook-writing helpers
# ============================================================
def safe_sheet_name(name, used_names):
    """Excel worksheet names: <=31 chars, no : \\ / ? * [ ], not blank,
    and must be unique within the workbook."""
    cleaned = re.sub(r"[:\\/?*\[\]]", "_", name).strip() or "Sheet"
    base = cleaned[:31]
    candidate = base
    n = 2
    while candidate.lower() in used_names:
        suffix = f" ({n})"
        candidate = base[: 31 - len(suffix)] + suffix
        n += 1
    used_names.add(candidate.lower())
    return candidate


def _to_num(v):
    """Convert an OHLC value to float where possible; leave untouched otherwise."""
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return v


_HEADER_FILL = PatternFill(start_color="E8EAF1", end_color="E8EAF1", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="1E1B2E")
_DATE_FMT = "yyyy-mm-dd"
_NUM_FMT = "#,##0.00"


def _style_header(ws, headers):
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"


def _write_index_sheet(ws, rows):
    """rows: list of dicts with date/open/high/low/close, already sorted by date."""
    _style_header(ws, ["Date", "Open", "High", "Low", "Close"])
    for r, row in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=row["date"]).number_format = _DATE_FMT
        for c, key in enumerate(("open", "high", "low", "close"), start=2):
            cell = ws.cell(row=r, column=c, value=row[key])
            if isinstance(row[key], (int, float)):
                cell.number_format = _NUM_FMT
    for i, w in enumerate([14, 13, 13, 13, 13], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_combined_sheet(ws, rows):
    """rows: list of dicts with index_name/date/open/high/low/close,
    already sorted by (date, index_name)."""
    _style_header(ws, ["Index", "Date", "Open", "High", "Low", "Close"])
    for r, row in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=row["index_name"])
        ws.cell(row=r, column=2, value=row["date"]).number_format = _DATE_FMT
        for c, key in enumerate(("open", "high", "low", "close"), start=3):
            cell = ws.cell(row=r, column=c, value=row[key])
            if isinstance(row[key], (int, float)):
                cell.number_format = _NUM_FMT
    for i, w in enumerate([22, 14, 13, 13, 13, 13], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
